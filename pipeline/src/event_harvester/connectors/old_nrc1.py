from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import load_workbook

from event_harvester.models import Event
from event_harvester.utils.hashing import stable_event_id


# NRC posts yearly FOIA Excel files like:
# https://nrc.uscg.mil/FOIAFiles/CY26.xlsx  (Calendar Year 2026)
NRC_FOIA_BASE = "https://nrc.uscg.mil/FOIAFiles"


# ---------- helpers ----------

def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _snake(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "unknown"


def _to_two_digit_year(y: int) -> str:
    return f"{y % 100:02d}"


def _nrc_year_url(year: int) -> str:
    yy = _to_two_digit_year(year)
    return f"{NRC_FOIA_BASE}/CY{yy}.xlsx"


def _parse_excel_datetime(v: Any) -> Optional[datetime]:
    """
    openpyxl may return datetimes directly, or strings, or empty cells.
    We try best-effort parsing.
    """
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)

    s = str(v).strip()
    if not s:
        return None

    # Try common patterns seen in exports
    for fmt in (
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass

    # Last resort: attempt ISO-ish parse
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def _normalize_header(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h).strip()).upper()


def _find_col(headers: List[str], candidates: List[str]) -> Optional[int]:
    """
    Find the first header index matching any candidate (case-insensitive, normalized).
    candidates are compared after uppercasing and collapsing spaces.
    """
    norm_headers = [_normalize_header(h) for h in headers]
    norm_candidates = [_normalize_header(c) for c in candidates]
    for cand in norm_candidates:
        for i, h in enumerate(norm_headers):
            if h == cand:
                return i
    # also allow partial match for robustness
    for cand in norm_candidates:
        for i, h in enumerate(norm_headers):
            if cand in h:
                return i
    return None


def _classify_event_type(raw: Dict[str, Any]) -> str:
    """
    Simple heuristic based on material/substance text.
    """
    text_parts = []
    for k in ("MATERIAL", "SUBSTANCE", "RELEASED", "DESCRIPTION", "INCIDENT DESCRIPTION", "INCIDENT_DESCRIPTION"):
        v = raw.get(k)
        if v:
            text_parts.append(str(v))
    text = " ".join(text_parts).lower()

    oil_keywords = [
        "oil", "diesel", "gasoline", "petroleum", "crude", "fuel",
        "jet fuel", "kerosene", "lubric", "hydraulic", "bunker",
    ]
    if any(k in text for k in oil_keywords):
        return "oil_spill"

    chem_keywords = [
        "chlorine", "ammonia", "acid", "caustic", "solvent", "benzene",
        "toluene", "xylene", "sulfur", "hydrogen", "propane", "butane",
        "ethylene", "methanol", "ethanol", "pesticide",
    ]
    if any(k in text for k in chem_keywords):
        return "chemical_release"

    return "industrial_incident"


# ---------- download + parse ----------

def download_nrc_year_excel(year: int, raw_dir: Path) -> Path:
    """
    Download NRC FOIA Excel for the given year if not present.
    Returns local filepath.
    """
    _ensure_dir(raw_dir)
    url = _nrc_year_url(year)
    out_path = raw_dir / f"NRC_CY{_to_two_digit_year(year)}.xlsx"

    if out_path.exists() and out_path.stat().st_size > 0:
        return out_path

    resp = requests.get(url, timeout=120)
    resp.raise_for_status()
    out_path.write_bytes(resp.content)
    
    # after writing out_path
    head = out_path.read_bytes()[:4]
    if head != b"PK\x03\x04":
        raise RuntimeError(
            f"Downloaded file is not a real .xlsx (expected ZIP header 'PK..'). "
            f"First 4 bytes: {head!r}. You may have downloaded an HTML error page."
        )
    return out_path


def iter_nrc_excel_rows(xlsx_path: Path) -> Iterable[Dict[str, Any]]:
    """
    Stream-ish parse using openpyxl. Yields dict per row.
    Assumes first row contains headers.
    """
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    # NRC files usually have the data in the first sheet
    ws = wb.worksheets[0]

    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return

    headers_list = [str(h).strip() if h is not None else "" for h in headers]

    for r in rows:
        if r is None:
            continue
        # skip totally empty rows
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        yield {headers_list[i]: r[i] for i in range(min(len(headers_list), len(r)))}
        



# ---------- normalization ----------

def row_to_event(raw: Dict[str, Any]) -> Optional[Event]:
    """
    Convert one NRC row dict to Event using the actual FOIA Excel headers you have:
      - SEQNOS (unique id)
      - DATE_TIME_RECEIVED / DATE_TIME_COMPLETE (timestamps)
      - RESPONSIBLE_STATE / RESPONSIBLE_CITY (location-ish fields)
      - SOURCE / CALLTYPE / RESPONSIBLE_COMPANY (context)
    """
    # Normalize keys once
    norm = {_normalize_header(k): v for k, v in raw.items()}

    # 1) Stable ID
    report_id_val = norm.get("SEQNOS")
    if report_id_val in (None, ""):
        return None
    report_id = str(report_id_val).strip()

    # 2) Datetime (prefer received)
    dt = _parse_excel_datetime(norm.get("DATE_TIME_RECEIVED")) or _parse_excel_datetime(norm.get("DATE_TIME_COMPLETE"))
    if not dt:
        return None

    # 3) State/county/city (use responsible fields as a start)
    state = norm.get("RESPONSIBLE_STATE")
    state = str(state).strip() if state not in (None, "") else None
    
    if state in {"XX", "NA", "N/A", "UNKNOWN"}:
        state = None


    county = norm.get("RESPONSIBLE_COUNTY")
    county = str(county).strip() if county not in (None, "") else None

    # (Optional)
    # Many NRC files have incident location columns too; if you find them later,
    # prefer incident-location over responsible-location.
    city = norm.get("RESPONSIBLE_CITY")
    city = str(city).strip() if city not in (None, "") else None

    # 4) Build a useful description from available fields
    calltype = norm.get("CALLTYPE")
    company = norm.get("RESPONSIBLE_COMPANY")
    source_field = norm.get("SOURCE")

    parts = []
    if calltype not in (None, ""):
        parts.append(f"CALLTYPE: {calltype}")
    if source_field not in (None, ""):
        parts.append(f"SOURCE: {source_field}")
    if company not in (None, ""):
        parts.append(f"RESPONSIBLE_COMPANY: {company}")
    if city:
        parts.append(f"CITY: {city}")
    if state:
        parts.append(f"STATE: {state}")

    description = " | ".join(parts) if parts else None

    # 5) Event typing (simple heuristic based on whatever fields are available)
    event_type = _classify_event_type(norm)  # will fall back to industrial_incident if no material fields

    source = "NRC"
    source_record_id = report_id
    event_id = stable_event_id(source, source_record_id)
    title = f"{event_type.replace('_', ' ').title()} - {state or 'NA'} - NRC {report_id}"

    # 6) Lat/Lon (if present later in the sheet; harmless if absent)
    lat = None
    lon = None
    for key in ("LATITUDE", "LAT"):
        if key in norm and norm[key] not in (None, ""):
            try:
                lat = float(norm[key])
            except Exception:
                lat = None
            break
    for key in ("LONGITUDE", "LON", "LONG"):
        if key in norm and norm[key] not in (None, ""):
            try:
                lon = float(norm[key])
            except Exception:
                lon = None
            break

    return Event(
        event_id=event_id,
        source=source,
        source_record_id=source_record_id,
        event_type=_snake(event_type),
        title=title,
        event_start=dt,
        event_end=None,
        state=state,
        county=county,
        lat=lat,
        lon=lon,
        description=description,
        url=None,
        raw=raw,
    )



# ---------- public API ----------

def bulk_events(year: int, raw_dir: Path) -> Iterable[Event]:
    """
    Bulk backfill for a full calendar year by downloading CY{YY}.xlsx and parsing all rows.
    """
    xlsx_path = download_nrc_year_excel(year, raw_dir=raw_dir)
    for raw in iter_nrc_excel_rows(xlsx_path):
        ev = row_to_event(raw)
        if ev is not None:
            yield ev


def incremental_events(
    start_date: date,
    end_date: date,
    raw_dir: Path,
    year: Optional[int] = None,
) -> Iterable[Event]:
    """
    Incremental updates by re-downloading the current year's file and filtering by event_start:
      start_date <= event_start.date() < end_date
    """
    y = year or date.today().year
    xlsx_path = download_nrc_year_excel(y, raw_dir=raw_dir)

    for raw in iter_nrc_excel_rows(xlsx_path):
        ev = row_to_event(raw)
        if ev is None:
            continue
        d = ev.event_start.date() if isinstance(ev.event_start, datetime) else None
        if d is None:
            continue
        if d < start_date:
            continue
        if d >= end_date:
            continue
        yield ev


def get_events_hybrid(
    bulk_years: List[int],
    raw_dir: Path,
    incremental_start: Optional[date] = None,
    incremental_end: Optional[date] = None,
) -> List[Event]:
    """
    Convenience function:
      - parse bulk years
      - optionally parse incremental window
    Dedup is enforced by event_id.
    """
    seen: set[str] = set()
    out: List[Event] = []

    for y in bulk_years:
        for ev in bulk_events(y, raw_dir=raw_dir / "bulk" / str(y)):
            if ev.event_id in seen:
                continue
            seen.add(ev.event_id)
            out.append(ev)

    if incremental_start and incremental_end:
        for ev in incremental_events(incremental_start, incremental_end, raw_dir=raw_dir / "incremental"):
            if ev.event_id in seen:
                continue
            seen.add(ev.event_id)
            out.append(ev)

    return out


if __name__ == "__main__":
    # Quick smoke test
    raw_dir = Path("data/raw/nrc")

    # Bulk: one year
    sample_year = 2020
    events = list(bulk_events(sample_year, raw_dir=raw_dir / "bulk" / str(sample_year)))
    print("Bulk events:", len(events))
    for e in events[:3]:
        print(e.model_dump())

    # Incremental: last 7 days of current year file
    today = date.today()
    start = today - timedelta(days=7)
    inc = list(incremental_events(start, today, raw_dir=raw_dir / "incremental"))

    print("Incremental events (last ~7 days):", len(inc))
    for e in inc[:3]:
        print(e.model_dump())
