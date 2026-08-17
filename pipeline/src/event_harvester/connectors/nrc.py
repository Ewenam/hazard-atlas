from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import requests
from openpyxl import load_workbook

from event_harvester.models import Event
from event_harvester.utils.hashing import stable_event_id


NRC_FOIA_BASE = "https://nrc.uscg.mil/FOIAFiles"


# ---------- helpers ----------

def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _snake(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s or "unknown"


def _to_two_digit_year(y: int) -> str:
    return f"{y % 100:02d}"


def _nrc_year_url(year: int) -> str:
    yy = _to_two_digit_year(year)
    return f"{NRC_FOIA_BASE}/CY{yy}.xlsx"


def _parse_excel_datetime(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)

    s = str(v).strip()
    if not s:
        return None

    for fmt in (
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass

    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def _normalize_header(h: Any) -> str:
    return re.sub(r"\s+", " ", str(h).strip()).upper()


def _as_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _as_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _as_float(v: Any) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _dms_to_decimal(deg: Optional[float], minute: Optional[float], sec: Optional[float], quad: Optional[str]) -> Optional[float]:
    if deg is None:
        return None
    m = minute or 0.0
    s = sec or 0.0
    val = abs(deg) + (m / 60.0) + (s / 3600.0)
    q = (quad or "").strip().upper()
    if q in {"S", "W"}:
        val = -val
    return val


# ---------- classification (improved) ----------

def _classify_event_type_from_text(text: str) -> str:
    t = (text or "").lower()

    oil_keywords = [
        "oil", "diesel", "gasoline", "petroleum", "crude", "fuel",
        "jet fuel", "kerosene", "lubric", "hydraulic", "bunker",
    ]
    if any(k in t for k in oil_keywords):
        return "oil_spill"

    chem_keywords = [
        "chlorine", "ammonia", "acid", "caustic", "solvent", "benzene",
        "toluene", "xylene", "sulfur", "hydrogen", "propane", "butane",
        "ethylene", "methanol", "ethanol", "pesticide",
    ]
    if any(k in t for k in chem_keywords):
        return "chemical_release"

    return "industrial_incident"


# ---------- download ----------

def download_nrc_year_excel(year: int, raw_dir: Path) -> Path:
    _ensure_dir(raw_dir)
    url = _nrc_year_url(year)
    out_path = raw_dir / f"NRC_CY{_to_two_digit_year(year)}.xlsx"

    if out_path.exists() and out_path.stat().st_size > 0:
        return out_path

    resp = requests.get(url, timeout=120)
    resp.raise_for_status()
    out_path.write_bytes(resp.content)

    head = out_path.read_bytes()[:4]
    if head != b"PK\x03\x04":
        raise RuntimeError(
            f"Downloaded file is not a real .xlsx (expected ZIP header 'PK..'). "
            f"First 4 bytes: {head!r}. You may have downloaded an HTML error page."
        )
    return out_path


# ---------- excel parsing ----------

def iter_nrc_sheet_rows(xlsx_path: Path, sheet_name: str) -> Iterable[Dict[str, Any]]:
    wb = load_workbook(filename=str(xlsx_path), read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)

    try:
        headers = next(rows)
    except StopIteration:
        return

    headers_list = [str(h).strip() if h is not None else "" for h in headers]

    for r in rows:
        if r is None:
            continue
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        yield {headers_list[i]: r[i] for i in range(min(len(headers_list), len(r)))}


# ---------- indexes for enrichment ----------

@dataclass(frozen=True)
class IncidentCommons:
    incident_dt: Optional[datetime]
    loc_state: Optional[str]
    loc_county: Optional[str]
    nearest_city: Optional[str]
    desc: Optional[str]
    type_of_incident: Optional[str]
    lat_deg: Optional[float]
    lat_min: Optional[float]
    lat_sec: Optional[float]
    lat_quad: Optional[str]
    lon_deg: Optional[float]
    lon_min: Optional[float]
    lon_sec: Optional[float]
    lon_quad: Optional[str]


def build_incident_commons_index(xlsx_path: Path) -> Dict[str, IncidentCommons]:
    idx: Dict[str, IncidentCommons] = {}
    for raw in iter_nrc_sheet_rows(xlsx_path, "INCIDENT_COMMONS"):
        norm = {_normalize_header(k): v for k, v in raw.items()}
        seq = norm.get("SEQNOS")
        if seq in (None, ""):
            continue
        seq = str(seq).strip()

        idx[seq] = IncidentCommons(
            incident_dt=_parse_excel_datetime(norm.get("INCIDENT_DATE_TIME")),
            loc_state=_as_str(norm.get("LOCATION_STATE")),
            loc_county=_as_str(norm.get("LOCATION_COUNTY")),
            nearest_city=_as_str(norm.get("LOCATION_NEAREST_CITY")),
            desc=_as_str(norm.get("DESCRIPTION_OF_INCIDENT")),
            type_of_incident=_as_str(norm.get("TYPE_OF_INCIDENT")),
            lat_deg=_as_float(norm.get("LAT_DEG")),
            lat_min=_as_float(norm.get("LAT_MIN")),
            lat_sec=_as_float(norm.get("LAT_SEC")),
            lat_quad=_as_str(norm.get("LAT_QUAD")),
            lon_deg=_as_float(norm.get("LONG_DEG")),
            lon_min=_as_float(norm.get("LONG_MIN")),
            lon_sec=_as_float(norm.get("LONG_SEC")),
            lon_quad=_as_str(norm.get("LONG_QUAD")),
        )
    return idx


@dataclass(frozen=True)
class IncidentDetails:
    any_injuries: Optional[str]
    number_injured: Optional[int]
    any_fatalities: Optional[str]
    number_fatalities: Optional[int]
    any_damages: Optional[str]
    damage_amount: Optional[float]


def build_incident_details_index(xlsx_path: Path) -> Dict[str, IncidentDetails]:
    """
    INCIDENT_DETAILS includes injuries/fatalities/damages/cost.
    """
    idx: Dict[str, IncidentDetails] = {}
    for raw in iter_nrc_sheet_rows(xlsx_path, "INCIDENT_DETAILS"):
        norm = {_normalize_header(k): v for k, v in raw.items()}
        seq = norm.get("SEQNOS")
        if seq in (None, ""):
            continue
        seq = str(seq).strip()

        idx[seq] = IncidentDetails(
            any_injuries=_as_str(norm.get("ANY_INJURIES")),
            number_injured=_as_int(norm.get("NUMBER_INJURED")),
            any_fatalities=_as_str(norm.get("ANY_FATALITIES")),
            number_fatalities=_as_int(norm.get("NUMBER_FATALITIES")),
            any_damages=_as_str(norm.get("ANY_DAMAGES")),
            damage_amount=_as_float(norm.get("DAMAGE_AMOUNT")),
        )
    return idx


def build_material_involved_index(xlsx_path: Path) -> Dict[str, List[Dict[str, Any]]]:
    """
    MATERIAL_INVOLVED can have multiple rows per SEQNOS.
    We store the raw material rows (small) so you can:
      - build classification text
      - keep it in Event.raw for traceability
    """
    idx: Dict[str, List[Dict[str, Any]]] = {}
    for raw in iter_nrc_sheet_rows(xlsx_path, "MATERIAL_INVOLVED"):
        norm = {_normalize_header(k): v for k, v in raw.items()}
        seq = norm.get("SEQNOS")
        if seq in (None, ""):
            continue
        seq = str(seq).strip()

        material_row = {
            "NAME_OF_MATERIAL": _as_str(norm.get("NAME_OF_MATERIAL")),
            "CHRIS_CODE": _as_str(norm.get("CHRIS_CODE")),
            "CAS_NUMBER": _as_str(norm.get("CAS_NUMBER")),
            "UN_NUMBER": _as_str(norm.get("UN_NUMBER")),
            "AMOUNT_OF_MATERIAL": norm.get("AMOUNT_OF_MATERIAL"),
            "UNIT_OF_MEASURE": _as_str(norm.get("UNIT_OF_MEASURE")),
            "IF_REACHED_WATER": _as_str(norm.get("IF_REACHED_WATER")),
            "AMOUNT_IN_WATER": norm.get("AMOUNT_IN_WATER"),
            "UNIT_OF_MEASURE_REACH_WATER": _as_str(norm.get("UNIT_OF_MEASURE_REACH_WATER")),
        }

        idx.setdefault(seq, []).append(material_row)

    return idx


def _materials_to_text(material_rows: List[Dict[str, Any]]) -> str:
    """
    Build a compact text blob for classification.
    """
    parts: List[str] = []
    for m in material_rows:
        name = m.get("NAME_OF_MATERIAL")
        cas = m.get("CAS_NUMBER")
        un = m.get("UN_NUMBER")
        chris = m.get("CHRIS_CODE")
        amt = m.get("AMOUNT_OF_MATERIAL")
        unit = m.get("UNIT_OF_MEASURE")
        reached_water = m.get("IF_REACHED_WATER")

        line_bits = []
        if name:
            line_bits.append(f"material {name}")
        if cas:
            line_bits.append(f"CAS {cas}")
        if un:
            line_bits.append(f"UN {un}")
        if chris:
            line_bits.append(f"CHRIS {chris}")
        if amt not in (None, ""):
            line_bits.append(f"amount {amt} {unit or ''}".strip())
        if reached_water:
            line_bits.append(f"reached_water {reached_water}")

        if line_bits:
            parts.append(" ".join(line_bits))
    return " | ".join(parts)


# ---------- event building (CALLS + enrichment) ----------

def row_to_event_enriched(
    calls_raw: Dict[str, Any],
    commons_idx: Dict[str, IncidentCommons],
    details_idx: Dict[str, IncidentDetails],
    materials_idx: Dict[str, List[Dict[str, Any]]],
) -> Optional[Event]:
    norm = {_normalize_header(k): v for k, v in calls_raw.items()}

    report_id_val = norm.get("SEQNOS")
    if report_id_val in (None, ""):
        return None
    report_id = str(report_id_val).strip()

    commons = commons_idx.get(report_id)
    details = details_idx.get(report_id)
    mats = materials_idx.get(report_id, [])

    # time: prefer incident datetime
    dt = (commons.incident_dt if commons else None) or _parse_excel_datetime(norm.get("DATE_TIME_RECEIVED")) or _parse_excel_datetime(norm.get("DATE_TIME_COMPLETE"))
    if not dt:
        return None

    # location: prefer incident location
    state = (commons.loc_state if commons else None) or norm.get("RESPONSIBLE_STATE")
    state = _as_str(state)
    if state and state.upper() in {"XX", "NA", "N/A", "UNKNOWN"}:
        state = None

    county = (commons.loc_county if commons else None)
    county = _as_str(county)

    city = (commons.nearest_city if commons else None) or norm.get("RESPONSIBLE_CITY")
    city = _as_str(city)

    # lat/lon: prefer DMS in commons
    lat = None
    lon = None
    if commons:
        lat = _dms_to_decimal(commons.lat_deg, commons.lat_min, commons.lat_sec, commons.lat_quad)
        lon = _dms_to_decimal(commons.lon_deg, commons.lon_min, commons.lon_sec, commons.lon_quad)

    # damages/injuries/fatalities
    injuries: Optional[int] = None
    fatalities: Optional[int] = None
    property_damage_usd: Optional[float] = None

    if details:
        # If explicitly "N", treat as 0; otherwise keep None if unknown
        if (details.any_injuries or "").upper() == "N":
            injuries = 0
        elif (details.any_injuries or "").upper() == "Y":
            injuries = details.number_injured if details.number_injured is not None else 0

        if (details.any_fatalities or "").upper() == "N":
            fatalities = 0
        elif (details.any_fatalities or "").upper() == "Y":
            fatalities = details.number_fatalities if details.number_fatalities is not None else 0

        if (details.any_damages or "").upper() == "N":
            property_damage_usd = 0.0
        elif (details.any_damages or "").upper() == "Y":
            property_damage_usd = details.damage_amount

    # classification text: commons description + type + materials
    cls_text_parts: List[str] = []
    if commons and commons.desc:
        cls_text_parts.append(commons.desc)
    if commons and commons.type_of_incident:
        cls_text_parts.append(commons.type_of_incident)
    if mats:
        cls_text_parts.append(_materials_to_text(mats))
    cls_text = " | ".join([p for p in cls_text_parts if p])

    event_type = _classify_event_type_from_text(cls_text)

    # description (human-readable)
    calltype = norm.get("CALLTYPE")
    company = norm.get("RESPONSIBLE_COMPANY")
    source_field = norm.get("SOURCE")

    parts = []
    if calltype not in (None, ""):
        parts.append(f"CALLTYPE: {calltype}")
    if source_field not in (None, ""):
        parts.append(f"SOURCE: {source_field}")
    if company not in (None, ""):
        parts.append(f"RESPONSIBLE_COMPANY: {company}")
    if commons and commons.type_of_incident:
        parts.append(f"INCIDENT_TYPE: {commons.type_of_incident}")
    if commons and commons.desc:
        parts.append(f"INCIDENT_DESC: {commons.desc}")
    if city:
        parts.append(f"CITY: {city}")
    if county:
        parts.append(f"COUNTY: {county}")
    if state:
        parts.append(f"STATE: {state}")
    if property_damage_usd is not None:
        parts.append(f"DAMAGE_USD: {property_damage_usd}")
    if injuries is not None:
        parts.append(f"INJURIES: {injuries}")
    if fatalities is not None:
        parts.append(f"FATALITIES: {fatalities}")

    description = " | ".join(parts) if parts else None

    source = "NRC"
    source_record_id = report_id
    event_id = stable_event_id(source, source_record_id)
    title = f"{event_type.replace('_', ' ').title()} - {state or 'NA'} - NRC {report_id}"

    raw_combined = {
        "CALLS": calls_raw,
        "INCIDENT_COMMONS": None if commons is None else {
            "INCIDENT_DATE_TIME": commons.incident_dt.isoformat() if commons.incident_dt else None,
            "LOCATION_STATE": commons.loc_state,
            "LOCATION_COUNTY": commons.loc_county,
            "LOCATION_NEAREST_CITY": commons.nearest_city,
            "TYPE_OF_INCIDENT": commons.type_of_incident,
            "DESCRIPTION_OF_INCIDENT": commons.desc,
            "LAT_DEG": commons.lat_deg, "LAT_MIN": commons.lat_min, "LAT_SEC": commons.lat_sec, "LAT_QUAD": commons.lat_quad,
            "LONG_DEG": commons.lon_deg, "LONG_MIN": commons.lon_min, "LONG_SEC": commons.lon_sec, "LONG_QUAD": commons.lon_quad,
        },
        "INCIDENT_DETAILS": None if details is None else {
            "ANY_INJURIES": details.any_injuries,
            "NUMBER_INJURED": details.number_injured,
            "ANY_FATALITIES": details.any_fatalities,
            "NUMBER_FATALITIES": details.number_fatalities,
            "ANY_DAMAGES": details.any_damages,
            "DAMAGE_AMOUNT": details.damage_amount,
        },
        "MATERIAL_INVOLVED": mats if mats else None,
        "CLASSIFIER_TEXT": cls_text or None,
    }

    return Event(
        event_id=event_id,
        source=source,
        source_record_id=source_record_id,
        event_type=_snake(event_type),
        title=title,
        event_start=dt,
        event_end=None,
        state=state,
        county=county,
        lat=lat,
        lon=lon,
        fatalities=fatalities,
        injuries=injuries,
        property_damage_usd=property_damage_usd,
        description=description,
        url=None,
        raw=raw_combined,
    )


# ---------- public API ----------

def bulk_events(year: int, raw_dir: Path) -> Iterable[Event]:
    xlsx_path = download_nrc_year_excel(year, raw_dir=raw_dir)

    commons_idx = build_incident_commons_index(xlsx_path)
    details_idx = build_incident_details_index(xlsx_path)
    materials_idx = build_material_involved_index(xlsx_path)

    for calls_raw in iter_nrc_sheet_rows(xlsx_path, "CALLS"):
        ev = row_to_event_enriched(calls_raw, commons_idx, details_idx, materials_idx)
        if ev is not None:
            yield ev


def incremental_events(
    start_date: date,
    end_date: date,
    raw_dir: Path,
    year: Optional[int] = None,
) -> Iterable[Event]:
    y = year or date.today().year
    xlsx_path = download_nrc_year_excel(y, raw_dir=raw_dir)

    commons_idx = build_incident_commons_index(xlsx_path)
    details_idx = build_incident_details_index(xlsx_path)
    materials_idx = build_material_involved_index(xlsx_path)

    for calls_raw in iter_nrc_sheet_rows(xlsx_path, "CALLS"):
        ev = row_to_event_enriched(calls_raw, commons_idx, details_idx, materials_idx)
        if ev is None:
            continue

        d = ev.event_start.date() if isinstance(ev.event_start, datetime) else None
        if d is None:
            continue
        if d < start_date:
            continue
        if d >= end_date:
            continue

        yield ev